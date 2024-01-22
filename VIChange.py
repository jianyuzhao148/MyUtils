import os
import sys

import openpyxl
import langid
from multiprocessing import Pool

import re


class String:
    def __init__(self):
        pass

    @staticmethod
    def extract_chinese(text: str) -> list:
        """
        获取字符串中中文
        :param text:
        :return:
        """
        pattern = re.compile(r'[\u4e00-\u9fff]+')
        chinese_chars = re.findall(pattern, text)
        return chinese_chars

    @staticmethod
    def str2unicode(text: str, encode="unicode_escape") -> bytes:
        return text.encode(encode)

    @staticmethod
    def unicode2str(b: bytes, encode="unicode_escape") -> str:
        return b.decode(encode)


class Log:
    def __init__(self):
        pass

    @staticmethod
    def ERROR(msg):
        print("\033[31m{0}\033[0m".format(msg))

    @staticmethod
    def INFO(msg):
        print("\033[32m{0}\033[0m".format(msg))

    @staticmethod
    def WARN(msg):
        print("\033[33m{0}\033[0m".format(msg))

    @staticmethod
    def DEBUG(msg):
        print("\033[34m{0}\033[0m".format(msg))


def over_call(p):
    Log.DEBUG("进程：" + str(os.getpid()) + "工作完成")


def error_call(error):
    print(f'进程:' + str(os.getpid()) + 'Error:', error, flush=True)


def write_sheet(excel_name, sheet_table: dict):
    number = 0
    Log.INFO("开始转换excel:{}".format(excel_name))
    workbook: openpyxl.workbook.Workbook = openpyxl.load_workbook(excel_name)
    for sheet_name in sheet_table.keys():
        Log.INFO("\t开始转换sheet:{}".format(sheet_name))
        sheet = workbook[sheet_name]
        local_table = sheet_table[sheet_name]
        for local in local_table.keys():
            sheet[local] = str(local_table[local])
            number = number + 1
    if number > 0:
        workbook.save(excel_name)
    workbook.close()


def get_all_sheet(excel_name):
    Log.INFO("开始读取{}".format(excel_name))
    sheet_table: dict = {}
    workbook: openpyxl.workbook.Workbook = openpyxl.load_workbook(excel_name)
    for sheet in workbook:
        sheet_table[sheet.title] = {}
        for i in sheet.iter_rows():
            for j in i:
                if type(j.value) == str and langid.classify(j.value)[0] == "vi":
                    Log.INFO("\t\t{}=>{}".format(j.value, String.str2unicode(j.value)))
                    sheet_table[sheet.title][j.coordinate] = String.str2unicode(j.value)
    workbook.close()
    return sheet_table


def restore_unicode(file_name):
    file_content = ""
    Log.INFO("\t{}".format(file_name))
    f = open(file_name, encoding="utf8")
    line = f.readline()
    while line:
        if "b'\"" in line:
            try:
                idx = line.index("b'\"")
                byte_str = line[idx + 2:-3]
                temp = eval("u" + "\'" + byte_str + "\'")
                try:
                    if "\'" in temp:
                        language_str = eval("u" + temp)
                        language_str = "\"" + language_str + "\""
                    else:
                        language_str = eval("u" + "\'" + temp + "\'")
                except:
                    language_str = eval("u" + temp)
                Log.INFO("\t\t{}=>{}".format(byte_str, language_str))
                line = line[:idx] + language_str + line[-2:]
            except:
                Log.ERROR(line)
        if "b'{" in line:
            idx = line.index("b'{")
            byte_str = line[idx + 1:-2]
            language_str = eval("u" + byte_str)
            language_str = eval("u" + "\'" + language_str + "\'")
            Log.INFO("\t\t{}=>{}".format(byte_str, language_str))
            line = line[:idx] + language_str + line[-2:]

        file_content = file_content + line
        line = f.readline()
    f.close()

    with open(file_name, "w", encoding="utf-8") as f:
        f.write(file_content)


def series_connection(file_path):
    sheet_table = get_all_sheet(file_path)
    write_sheet(file_path, sheet_table)


export_dict = {}


def get_file_quotation(file_path):
    symbol_1 = "\""
    symbol_2 = "\'"
    symbol_3 = "`"
    """
    获取文件中的引号文本
    :param file_path: 文件路径
    :return:
    """

    def method(symbol):
        global export_dict
        split_text_list = line.split(symbol)
        for i in range(len(split_text_list)):
            split_text = split_text_list[i]
            if langid.classify(split_text)[0] == "zh":
                Log.INFO("\t{}".format(split_text))
                if not export_dict.get(split_text):
                    export_dict[split_text] = []
                export_dict[split_text].append("{}@{}@{}@{}".format(file_path, line_number, i, symbol))

    with open(file_path, encoding="utf8") as f:
        line_number = 1
        for line in f.readlines():
            # 去除注释
            if "//" in line:
                line = line[: line.index("//")]
            if symbol_1 in line:
                method(symbol_1)
            if symbol_2 in line:
                method(symbol_2)
            if symbol_3 in line:
                method(symbol_3)
            line_number = line_number + 1


def write2excel(excel_name, sheet_name):
    workbook: openpyxl.workbook.Workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(sheet_name)
    row = 1
    for i in export_dict.keys():
        worksheet.cell(row, 1).value = i
        worksheet.cell(row, 2).value = len(export_dict[i])
        col = 3
        for j in export_dict[i]:
            worksheet.cell(row, col).value = j
            col = col + 1
        row = row + 1
    workbook.save(excel_name)


def read_excel(excel_name, sheet_name):
    translate_column = 1
    num_column = 2

    file_change_dict = {}
    workbook: openpyxl.workbook.Workbook = openpyxl.load_workbook(excel_name)
    worksheet = workbook[sheet_name]
    for i in worksheet.iter_rows():
        for k in range(1, i[num_column].value + 1):
            value: str = i[num_column + k].value
            file_name: str = value.split("@")[0]
            if not file_change_dict.get(file_name):
                file_change_dict[file_name] = []
            file_change_dict[file_name].append("{}@{}".format(value, i[translate_column].value))
    return file_change_dict


def write_translate_file(file_change_dict: dict):
    def write2file_by_line(file_path, line_num, index, symbol, value):
        Log.INFO("写入:" + value)
        lines = ""
        with open(file_path, encoding="utf8") as f:
            count_num = 1
            for line in f.readlines():
                if count_num == int(line_num):
                    try:
                        str_arr = line.split(symbol)
                        start = line.index(str_arr[int(index)])
                        end = start + len(str_arr[int(index)])
                        line = line[:start] + value + line[end:]
                    except:
                        print("a")
                lines = lines + line
                count_num = count_num + 1
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(lines)

    for file_path in file_change_dict.keys():
        file_change_list: list = file_change_dict[file_path]
        for i in file_change_list:
            temp = i.split("@")
            write2file_by_line(temp[0], temp[1], temp[2], temp[3], temp[4])


if __name__ == "__main__":
    str_help = """
    -h                      : 弹出帮助界面
    -vi2Unicode             : 转换越南文为Unicode字符,path为目录路径
    -revertUnicode          : 还原Unicode
    -extractChinese         : 文件中提取中文
    -revertExtractChinese   : 还原提取的中文到文件
    """
    # 参数识别
    if len(sys.argv) <= 1:
        Log.INFO(str_help)
    else:
        param = sys.argv[1]
        if param == "-h":
            Log.INFO(str_help)
        elif param == "-vi2Unicode":
            progress_pool = Pool()
            for root, dirs, files in os.walk("表格导出"):
                for file in files:
                    # 提供文件名，目录忽略
                    if file.endswith("xlsx") and (not (file in ["D-掉落表.xlsx"])) and (
                            not (root.split("\\")[-1] in ["测试表"])):
                        file_path = os.path.join(root, file)
                        progress_pool.apply_async(series_connection, args=(file_path,), callback=over_call,
                                                  error_callback=error_call)

            progress_pool.close()
            progress_pool.join()
        elif param == "-revertUnicode":
            for root, dirs, files in os.walk("client\\data"):
                for file in files:
                    if file.endswith("config") or file.endswith("confg"):
                        restore_unicode(os.path.join(root, file))
        elif param == "-extractChinese":
            for root, dirs, files in os.walk(r"..\..\client\main\src"):
                for file in files:
                    file_path = os.path.join(root, file)
                    Log.INFO("开始读取{}".format(file))
                    get_file_quotation(file_path)
            write2excel(r"./chinese.xlsx", "chinese")
        elif param == "-revertExtractChinese":
            d = read_excel(r"D:\LOVietnamese\config\packet\chinese.xlsx", "chinese")
            write_translate_file(d)
