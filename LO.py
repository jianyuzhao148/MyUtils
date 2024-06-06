import os
import re
import sys
from multiprocessing import Pool

import langid
import openpyxl
from PyQt5.QtCore import QStringListModel, Qt, QThread
from PyQt5.QtWidgets import (QAbstractItemView, QApplication, QCompleter,
                             QGroupBox, QHBoxLayout, QLabel, QLineEdit,
                             QListView, QMainWindow, QMessageBox, QProgressBar,
                             QPushButton, QVBoxLayout)

include_file_path = r"D:\LoVN\config\packet\表格导出"
include_file_last = ".xlsx"
revert_path = r"D:\LoVN\config\packet\client\data"


class String:
    def __init__(self):
        pass

    @staticmethod
    def extract_chinese(text: str) -> list:
        """
        获取字符串中中文
        :param text:
        :return:
        """
        pattern = re.compile(r"[\u4e00-\u9fff]+")
        chinese_chars = re.findall(pattern, text)
        return chinese_chars

    @staticmethod
    def str2unicode(text: str, encode="unicode_escape") -> bytes:
        return text.encode(encode)

    @staticmethod
    def unicode2str(b: bytes, encode="unicode_escape") -> str:
        return b.decode(encode)


def get_all_sheet(excel_name):
    sheet_table: dict = {}
    workbook: openpyxl.Workbook = openpyxl.load_workbook(excel_name)
    for sheet in workbook:
        sheet_table[sheet.title] = {}
        for i in sheet.iter_rows():
            for j in i:
                if type(j.value) == str and langid.classify(j.value)[0] == "vi":
                    sheet_table[sheet.title][j.coordinate] = String.str2unicode(j.value)
    workbook.close()
    return sheet_table


def series_connection(file_path):
    sheet_table = get_all_sheet(file_path)
    write_sheet(file_path, sheet_table)


def write_sheet(excel_name, sheet_table: dict):
    number = 0
    workbook: openpyxl.Workbook = openpyxl.load_workbook(excel_name)
    for sheet_name in sheet_table.keys():
        sheet = workbook[sheet_name]
        local_table = sheet_table[sheet_name]
        for local in local_table.keys():
            sheet[local] = str(local_table[local])
            number = number + 1
    if number > 0:
        workbook.save(excel_name)
    workbook.close()


class MainQWidget(QMainWindow):

    def __restore_unicode(self, file_name):
        file_content = ""
        f = open(file_name, encoding="utf8")
        line = f.readline()
        while line:
            if "b'\"" in line:
                try:
                    idx = line.index("b'\"")
                    byte_str = line[idx + 2 : -3]
                    temp = eval("u" + "'" + byte_str + "'")
                    try:
                        if "'" in temp:
                            language_str = eval("u" + temp)
                            language_str = '"' + language_str + '"'
                        else:
                            language_str = eval("u" + "'" + temp + "'")
                    except:
                        language_str = eval("u" + temp)
                    line = line[:idx] + language_str + line[-2:]
                except:
                    pass
            if "b'{" in line:
                idx = line.index("b'{")
                byte_str = line[idx + 1 : -2]
                language_str = eval("u" + byte_str)
                language_str = eval("u" + "'" + language_str + "'")
                line = line[:idx] + language_str + line[-2:]

            file_content = file_content + line
            line = f.readline()
        f.close()

        with open(file_name, "w", encoding="utf-8") as f:
            f.write(file_content)

    def __complet_call(self, _):
        self.file_len = self.file_len - 1
        self.progress.setValue(1)

    def __transform_vi2unicode(self, file_list):
        self.file_len = len(file_list)
        self.progress.setMaximum(self.file_len)
        progress_pool = Pool()
        for file in file_list:
            progress_pool.apply_async(
                series_connection,
                args=(file,),
                callback=self.__complet_call,
                error_callback=None,
            )
        progress_pool.close()
        progress_pool.join()

    def __transform(self):
        select_list = self.listview.selectedIndexes()
        if len(select_list) == 0:
            QMessageBox.warning(
                self,
                "警告",
                "未选择任何文件",
                QMessageBox.Yes,
            )
            return
        select_file_str = ""
        file_list = []
        for i in select_list:
            select_file_str = select_file_str + "\n" + i.data()
            file_list.append(include_file_path + os.sep + i.data())

        reply = QMessageBox.information(
            self,
            "Tips",
            "请核对选中文件：" + select_file_str,
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply == QMessageBox.Yes:
            self.__transform_vi2unicode(file_list)
        QMessageBox.information(
            self,
            "提示",
            "转换完成",
            QMessageBox.Yes,
        )

    def __translate_all(self):
        result = QMessageBox.warning(
            self,
            "警告",
            "全部转换需要较长时间,是否继续?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if result == QMessageBox.Yes:
            file_list = []
            for i in self.file_list:
                file_list.append(include_file_path + os.sep + i)
            self.__transform_vi2unicode(file_list)
            QMessageBox.information(
                self,
                "提示",
                "转换完成",
                QMessageBox.Yes,
            )

    def __check_excel(self):
        all_items = os.listdir(include_file_path)
        self.file_list = [
            f
            for f in all_items
            if (
                os.path.isfile(os.path.join(include_file_path, f))
                and os.path.splitext(f)[-1] == include_file_last
            )
        ]

    def __revert_unicode(self):
        for root, _, files in os.walk(revert_path):
            for file in files:
                if file.endswith("config") or file.endswith("confg"):
                    self.__restore_unicode(os.path.join(root, file))
        QMessageBox.information(
            self,
            "提示",
            "转换完成",
            QMessageBox.Yes,
        )

    def __start_outsheet(self):
        os.system(include_file_path + os.sep + "导表管理器.exe")

    def search_lineEdit_finished(self):
        print("失去焦点")

    def __init__(self, parent=None):
        super().__init__(parent)
        self.file_list = []
        self.file_len = 0
        self.__check_excel()
        self.setAttribute(Qt.WA_TranslucentBackground)
        # self.setWindowFlags(Qt.WindowStaysOnTopHint|Qt.FramelessWindowHint|Qt.Tool);

        self.setWindowTitle("文本转换工具")
        self.setFixedSize(600, 500)

        group_box = QGroupBox("文本转换", self)

        self.parent_layout = QVBoxLayout()
        group_box.setLayout(self.parent_layout)

        self.select_layout = QVBoxLayout()

        # self.search_layout = QHBoxLayout()
        # self.search_label = QLabel("搜索:")
        # self.search_completer = QCompleter(self.file_list)
        # self.search_completer.setFilterMode(Qt.MatchFlag.MatchContains)
        # self.search_lineEdit = QLineEdit()
        # self.search_lineEdit.editingFinished.connect(self.search_lineEdit_finished)
        # self.search_lineEdit.setCompleter(self.search_completer)
        # self.search_layout.addWidget(self.search_label)
        # self.search_layout.addWidget(self.search_lineEdit)
        # self.select_layout.addLayout(self.search_layout)

        self.listview = QListView()
        self.string_list = QStringListModel()
        self.string_list.setStringList(self.file_list)
        self.listview.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.listview.setModel(self.string_list)
        self.select_layout.addWidget(self.listview)

        self.opation_layout = QVBoxLayout()

        self.progress_layout = QHBoxLayout()
        self.progress_label = QLabel("进度:")
        self.progress = QProgressBar()
        self.progress_layout.addWidget(self.progress_label)
        self.progress_layout.addWidget(self.progress)
        self.opation_layout.addLayout(self.progress_layout)

        self.transform_layout = QHBoxLayout()
        self.transform_label = QLabel("越文转Unicode:")
        self.transform_button = QPushButton("开始转换")
        self.transform_all_button = QPushButton("全部转换")
        self.transform_button.clicked.connect(self.__transform)
        self.transform_all_button.clicked.connect(self.__translate_all)
        self.transform_layout.addWidget(self.transform_label)
        self.transform_layout.addWidget(self.transform_button)
        self.transform_layout.addWidget(self.transform_all_button)
        self.opation_layout.addLayout(self.transform_layout)

        self.outsheet_layout = QHBoxLayout()
        self.outsheet_button = QPushButton("启动导表工具")
        self.outsheet_button.clicked.connect(self.__start_outsheet)
        self.outsheet_layout.addWidget(self.outsheet_button)
        self.opation_layout.addLayout(self.outsheet_layout)

        self.revert_layout = QHBoxLayout()
        self.revert_button = QPushButton("Unicode恢复")
        self.revert_button.clicked.connect(self.__revert_unicode)
        self.revert_layout.addWidget(self.revert_button)
        self.opation_layout.addLayout(self.revert_layout)

        self.parent_layout.addLayout(self.select_layout)
        self.parent_layout.addLayout(self.opation_layout)

        self.setCentralWidget(group_box)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    dlg = MainQWidget()
    dlg.show()
    sys.exit(app.exec_())
