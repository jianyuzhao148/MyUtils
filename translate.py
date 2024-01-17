import os
import re
import sys
import openpyxl

import Util.String
from Util.File import File


class Item:
    """结构体"""
    def __init__(self, path, line, start, end, value):
        self.path = path
        self.line = line
        self.start = start
        self.end = end
        self.value = value

    def get_list(self):
        return [self.path, self.line, self.start, self.end, self.value]

    def description(self):
        return "{},{},{},{},{}".format(self.path, self.line, self.start, self.end, self.value)


items: dict[str, list] = {}  # {str:list}


def extract_chinese_characters(file_path):
    """通过提取出双引号中的字符判断其中有没有中文字符做判断"""
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
        line_number = 1
        for line in file.readlines():
            for m in re.finditer(r'"([^"]*)"', line):
                text = m.group(1)
                if len(Util.String.extract_chinese(text)) > 0:
                    if not items.get(text):
                        items[text] = []
                    items.get(text).append(Item(file_path, line_number, m.start() + 1, m.end(), text))
            line_number = line_number + 1


def traverse_folder(root_folder, file_last):
    """遍历文件夹下所有文件"""
    for root, dirs, files in os.walk(root_folder):
        for fileObj in files:
            if fileObj.split(".")[-1] == file_last:
                extract_chinese_characters(os.path.join(root, fileObj))


def write_excel(excel_name, sheet_name):
    """写入excel"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(sheet_name)
    row = 1
    for i in items.keys():
        col: int = 2
        worksheet.cell(row, col, i)
        for j in items.get(i):
            col = col + 1
            worksheet.cell(row, col, j.description())
        row = row + 1
    workbook.save(excel_name)


def read_excel(excel_name, sheet_name):
    """读取excel """
    workbook: openpyxl.workbook.Workbook = openpyxl.load_workbook(excel_name)
    sheet = workbook[sheet_name]
    for i in sheet.iter_rows():
        i_old = i
        i = i[2:]
        for j in i:
            # 忽略空cell
            if not j.value:
                break
            File.replace(j.value.split(",")[0],i_old[1].value,i_old[0].value)


if __name__ == "__main__":
    assert sys.version_info[0] == 3, "当前环境为Python{}.{},请安装Python3后重试".format(sys.version_info[0],
                                                                                        sys.version_info[1])

    traverse_folder("D:\\LOVietnamese\\client\\main\\resource\\skins\\shengshou", "exml")
    write_excel("Test.xlsx", "test")
    read_excel("Test.xlsx", "test")