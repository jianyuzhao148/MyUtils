import json
import os
import re
import sys

import openpyxl
import requests
from PyQt5.QtWidgets import (QApplication, QFileDialog, QGroupBox, QHBoxLayout,
                             QLabel, QLineEdit, QMainWindow, QMessageBox,
                             QProgressBar, QPushButton, QVBoxLayout)


def extract_chinese(text: str) -> list:
    """
    获取字符串中中文
    :param text:
    :return:
    """
    pattern = re.compile(r"[\u4e00-\u9fff]+")
    chinese_chars = re.findall(pattern, text)
    return chinese_chars


def remove_space(str) -> str:
    """
    去除空格
    :param str:
    :return:
    """
    return "".join(str.split())


def ERROR(msg):
    print("\033[31m{0}\033[0m".format(msg))


"""
究极翻译脚本
"""


class MainQWidget(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.put_xlsx_path = ""
        self.put_translate_xlsx_path = ""
        self.put_path = ""
        self.include_last = [".m", ".mm"]

        self.__sheet_name = "word"
        self.__out_path = "Extract.xlsx"

        self.setWindowTitle("文字提取工具-测试版")
        self.root_box = QVBoxLayout()
        group_box = QGroupBox("文字提取", self)
        self.write_layout = QHBoxLayout()
        self.read_layout = QHBoxLayout()
        self.translate_layout = QHBoxLayout()
        self.progress_layout = QHBoxLayout()
        group_box.setLayout(self.root_box)
        self.root_box.addLayout(self.read_layout)
        self.root_box.addLayout(self.write_layout)
        self.root_box.addLayout(self.translate_layout)
        self.root_box.addLayout(self.progress_layout)

        self.label_write = QLabel()
        self.label_write.setText("写入文字:")
        self.label_read = QLabel()
        self.label_read.setText("提取文字:")
        self.label_translate = QLabel()
        self.label_translate.setText("AI  翻译:")
        self.qlineEdit_selectfile = QLineEdit("Excel路径")
        self.qlineEdit_selectfile.setReadOnly(True)
        self.qpush_selectfile = QPushButton("选择Excel")

        self.write_layout.addWidget(self.label_write)
        self.write_layout.addWidget(self.qlineEdit_selectfile)
        self.write_layout.addWidget(self.qpush_selectfile)
        self.write_btn = QPushButton("开始写入")
        self.write_layout.addWidget(self.write_btn)
        self.qlineEdit_selectfolder = QLineEdit("目录路径")
        self.qlineEdit_selectfolder.setReadOnly(True)
        self.qpush_selectfolder = QPushButton("选择目录")
        self.read_btn = QPushButton("开始提取")
        self.read_layout.addWidget(self.label_read)
        self.read_layout.addWidget(self.qlineEdit_selectfolder)
        self.read_layout.addWidget(self.qpush_selectfolder)
        self.read_layout.addWidget(self.read_btn)

        self.translate_selectfile = QLineEdit("Excel路径")
        self.translate_selectfile.setReadOnly(True)
        self.qpush_translate_selectfile = QPushButton("选择Excel")
        self.translate_btn = QPushButton("开始翻译")
        self.translate_layout.addWidget(self.label_translate)
        self.translate_layout.addWidget(self.translate_selectfile)
        self.translate_layout.addWidget(self.qpush_translate_selectfile)
        self.translate_layout.addWidget(self.translate_btn)

        self.label_load = QLabel()
        self.label_load.setText("进度:")
        self.progress = QProgressBar()
        self.progress_layout.addWidget(self.label_load)
        self.progress_layout.addWidget(self.progress)

        self.qpush_selectfile.clicked.connect(self.select_file)
        self.qpush_translate_selectfile.clicked.connect(self.select_translate_file)
        self.qpush_selectfolder.clicked.connect(self.select_folder)
        self.read_btn.clicked.connect(self.extract)
        self.write_btn.clicked.connect(self.write)
        self.translate_btn.clicked.connect(self.translate)

        self.setCentralWidget(group_box)

    def translate(self):
        self.progress.setValue(0)
        if not os.path.exists(self.put_translate_xlsx_path):
            QMessageBox.warning(
                self,
                "警告",
                "无效路径",
                QMessageBox.Yes,
            )
            return
        # 读取需要翻译的字段
        excel_content = read_excel_col(
            self.put_translate_xlsx_path, self.__sheet_name, 1
        )
        translate_list = kimi_ask(excel_content)
        self.progress.setMaximum(len(translate_list))
        write_word_col(
            self.put_translate_xlsx_path,
            self.__sheet_name,
            translate_list,
            lambda i: self.progress.setValue(i + 1),
        )

    def select_file(self):
        self.put_xlsx_path, _ = QFileDialog.getOpenFileName(
            self,
            "getOpenFileName",
            "./",
            "Excel(*.xlsx);;",
        )
        self.qlineEdit_selectfile.setText(self.put_xlsx_path)

    def select_translate_file(self):
        self.put_translate_xlsx_path, _ = QFileDialog.getOpenFileName(
            self,
            "getOpenFileName",
            "./",
            "Excel(*.xlsx);;",
        )
        self.translate_selectfile.setText(self.put_translate_xlsx_path)

    def select_folder(self):
        self.put_path = QFileDialog.getExistingDirectory(self, "Select Directory", "./")
        self.qlineEdit_selectfolder.setText(self.put_path)

    def extract(self):
        self.progress.setValue(0)
        if not os.path.exists(self.put_path):
            QMessageBox.warning(
                self,
                "警告",
                "无效路径",
                QMessageBox.Yes,
            )
            return
        dict_file = {}
        for root, _, files in os.walk(self.put_path):
            for file in files:
                if os.path.splitext(file)[-1] in self.include_last:
                    file_path = os.path.join(root, file)
                    dict_temp = read_word(file_path=file_path)
                    if len(dict_temp) > 1:
                        dict_file[file_path] = dict_temp
        self.progress.setMaximum(len(dict_file.keys()))

        # 写入excel
        write2excel(
            self.__out_path,
            self.__sheet_name,
            dict_file,
            call_back=lambda i: self.progress.setValue(i),
        )
        result = QMessageBox.information(
            self, "提示", "成功,是否打开文件？", QMessageBox.Yes, QMessageBox.No
        )
        if result:
            os.startfile("Extract.xlsx")

    def write(self):
        self.progress.setValue(0)
        if not os.path.exists(self.put_xlsx_path):
            QMessageBox.warning(
                self,
                "警告",
                "无效路径",
                QMessageBox.Yes,
            )
            return
        dict_file = read_excel(self.put_xlsx_path, self.__sheet_name)
        self.progress.setMaximum(len(dict_file.keys()))
        count = 0
        for i in dict_file.keys():
            replace2file(i, dict_file[i])
            count = count + 1
            self.progress.setValue(count)
        QMessageBox.information(
            self,
            "提示",
            "成功",
            QMessageBox.Yes,
        )


def replace2file(file_path, line_word_map: dict) -> bool:
    """
    指定格式写入文件
    :param file_path:
    :param line_word_map:
    line_word_map形如:
    {
        line_number:[[old_word_list],[new_word_list]]
    }
    :return: bool
    """
    lines = ""
    with open(file_path, encoding="utf8") as f:
        line_number = 0
        for line in f.readlines():
            line_number = line_number + 1
            if line_number in line_word_map.keys():
                line_word_list: list = line_word_map[line_number]
                old_word_list = line_word_list[0]
                new_word_list = line_word_list[1]
                if len(old_word_list) == len(new_word_list):
                    for index in range(len(old_word_list)):
                        old_word = str(old_word_list[index])
                        new_word = str(new_word_list[index])

                        old_word_len = len(old_word)
                        start_index = line.find(old_word)
                        if start_index != -1:
                            end_index = start_index + old_word_len
                            # 拼接新line
                            line = line[:start_index] + new_word + line[end_index:]
                        else:
                            ERROR(
                                "File:{} Line:{}\n\t Info:{} not this line".format(
                                    file_path, line_number, old_word
                                )
                            )
                            return False
            lines = lines + line
    with open(file_path, "w", encoding="utf8") as f:
        f.write(lines)
    return True


def read_word(file_path) -> dict:
    """
    读取文字
    :param file_path:
    :return:dict 形如:
    {
        line_number:[word,...]
    }
    """
    with open(file_path, encoding="utf8") as f:
        dict_word = {}
        in_comment = False
        line_number = 0
        for line in f.readlines():
            line_number = line_number + 1
            # 处理注释
            if "/*" in line and "*/" in line:
                continue
            if "/**" in line and "*/" in line:
                continue
            if "/**" in remove_space(line):
                in_comment = True
            if in_comment and remove_space(line) == "*/":
                in_comment = False
            if "/*" in remove_space(line):
                in_comment = True
            if in_comment and remove_space(line) == "*/":
                in_comment = False
            if in_comment:
                continue
            if "//" in line:
                line = line[: line.index("//")]

            str_list = extract_chinese(line)
            if len(str_list) > 0:
                dict_word[line_number] = str_list
                # print("\n\tline: {} {}".format(line_number, str_list))
        return dict_word


def write2excel(excel_name, sheet_name, dict_file, call_back):
    workbook: openpyxl.Workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(sheet_name)
    row = 1
    count = 0
    for file_name in dict_file.keys():
        for file_line in dict_file[file_name].keys():
            for word in dict_file[file_name][file_line]:
                worksheet.cell(row, 1).value = file_name
                worksheet.cell(row, 2).value = file_line
                worksheet.cell(row, 3).value = word
                row = row + 1
        count = count + 1
        call_back(count)
    workbook.save(excel_name)


def read_excel(excel_name, sheet_name) -> dict:
    workbook: openpyxl.Workbook = openpyxl.load_workbook(excel_name)
    worksheet = workbook[sheet_name]
    dict_file = {}
    for i in worksheet.iter_rows():
        file_path = i[0].value
        line_number = i[1].value
        old_word = i[2].value
        new_word = i[3].value
        if not dict_file.get(file_path):
            dict_file[file_path] = {}
        if not dict_file[file_path].get(line_number):
            dict_file[file_path][line_number] = []
            dict_file[file_path][line_number].append([])
            dict_file[file_path][line_number].append([])
        dict_file[file_path][line_number][0].append(old_word)
        dict_file[file_path][line_number][1].append(new_word)
    return dict_file


def read_excel_col(excel_name, sheet_name, cole_num) -> list:
    workbook: openpyxl.Workbook = openpyxl.load_workbook(excel_name)
    worksheet: openpyxl.Workbook = workbook[sheet_name]
    list = []
    temp = 0
    for i in worksheet.iter_cols():
        if temp == cole_num:
            for j in i:
                list.append(j.value)
        temp = temp + 1
    return list


def kimi_ask(word_list) -> list:
    translate_list = word_list
    temp_str = ""
    for i in range(0, len(word_list)):
        temp_str = temp_str + "|" + (word_list[i] if word_list[i] is not None else "None")
        if i % 10 == 0:
            temp_str = temp_str[1:]
            translate_word = ask_translate(temp_str)
            sp_list = (translate_word.split("|"))
            sp_list.reverse()
            for j in range(0, len(sp_list)):
                translate_list[i-j] = sp_list[j]
            temp_str = ""
    return translate_list


def write_word_col(excel_name, sheet_name, word_list, call_back):
    workbook: openpyxl.Workbook = openpyxl.load_workbook(excel_name)
    worksheet = workbook[sheet_name]
    for i in range(0, len(word_list)):
        worksheet.cell(i + 1, 4).value = word_list[i]
        call_back(i)
    workbook.save(excel_name)


# Config
kimi_start: str = 'pm2 start dist/index.js --name "kimi-free-api"'
# domain: str = "http://192.168.31.130:8000"
domain: str = "http://127.0.0.1:8000"
token: str = (
    "eyJhbGciOiJIUzUxMiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJ1c2VyLWNlbnRlciIsImV4cCI6MTczNjc2NjA4MSwiaWF0IjoxNzI4OTkwMDgxLCJqdGkiOiJjczc0bjBhbjZ2dWttZ3Uzbjh2MCIsInR5cCI6InJlZnJlc2giLCJhcHBfaWQiOiJraW1pIiwic3ViIjoiY3M3NG4wYW42dnVrbWd1M244dTAiLCJzcGFjZV9pZCI6ImNzNzRuMGFuNnZ1a21ndTNuOHRnIiwiYWJzdHJhY3RfdXNlcl9pZCI6ImNzNzRuMGFuNnZ1a21ndTNuOHQwIiwiZGV2aWNlX2lkIjoiNzM0ODY4ODc5ODY0Mjg5MDc2MSJ9.U4prBShZCQa-iz2XU37j4fKdjnjXzzQAXJwpe1LV-XDUwzVHLj-q_rrHhGsmg9-LnGA70_82LOOg--x_lWnVCg")


def start_kimi_server():
    os.system(kimi_start)


def ask_translate(text: str) -> str:
    message = "将中文:'{}'翻译为泰文，无法翻译的请意译，必须只返回翻译文本，不要返回翻译来源,只要翻译的文本！！！".format(
        text)
    content = ask(message)
    # 如果返回的数据长度不对则重新翻译
    if len(message.split("|")) != len(content.split("|")):
        return ask_translate(text)
    return content


# 提问
def ask(message, use_search: bool = True) -> str:
    url = "{}/v1/chat/completions".format(domain)
    json_data = json.dumps(
        {
            "model": "kimi",
            "messages": [{"role": "user", "content": message}],
            "use_search": use_search,
            "stream": False,
        }
    )
    headers = {
        "Content-Type": "application/json",
        "Authorization": "Bearer {}".format(token),
    }
    try:
        response = requests.request("POST", url, headers=headers, data=json_data)
        msg = response.content.decode()
        content = json.loads(msg)["choices"][0]["message"]["content"]
        return content
    except Exception:
        print("ERROR:" + msg)
    return ""


if __name__ == "__main__":
    app = QApplication(sys.argv)
    dlg = MainQWidget()
    dlg.show()
    sys.exit(app.exec_())

    #     # 单文件提取
    #     with open(param, encoding="utf8") as f:
    #         print(read_word(file_path=param))
